from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import date
import pandas as pd
import io
from typing import Optional
from ..models.cita import Cita
from ..models.cliente import Cliente
from ..models.especialista import Especialista
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

class ReportService:
    @staticmethod
    def export_citas_excel(
        db: Session,
        sede_id: int,
        fecha_inicio: date,
        fecha_fin: date,
        incluir_cliente: bool = False
    ) -> io.BytesIO:
        """
        Genera un archivo Excel con el reporte de citas.
        """
        query = db.query(Cita).filter(
            Cita.sede_id == sede_id,
            Cita.fecha >= fecha_inicio,
            Cita.fecha <= fecha_fin
        )
        
        citas = query.order_by(Cita.fecha, Cita.hora_inicio).all()
        
        data = []
        for cita in citas:
            row = {
                "Fecha": cita.fecha.strftime("%Y-%m-%d"),
                "Hora": cita.hora_inicio.strftime("%H:%M"),
                "Especialista": f"{cita.especialista.nombre} {cita.especialista.apellido or ''}".strip(),
                "Servicio": cita.servicio.nombre,
            }
            
            if incluir_cliente:
                if cita.cliente:
                    row["Cliente"] = f"{cita.cliente.nombre} {cita.cliente.apellido or ''}".strip()
                    row["Teléfono"] = cita.cliente.telefono or "N/A"
                    row["Correo"] = cita.cliente.email or "N/A"
                else:
                    row["Cliente"] = "Anónimo"
                    row["Teléfono"] = "N/A"
                    row["Correo"] = "N/A"
            
            data.append(row)
            
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Citas')
            
            worksheet = writer.sheets['Citas']
            
            # Definir estilos
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark blue
            header_font = Font(color="FFFFFF", bold=True)
            row_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Light blue (azulito)
            
            border_side = Side(style='thin', color="000000")
            thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            
            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Aplicar estilos a las filas de datos
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left')

            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns):
                max_len = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = max_len
                
        output.seek(0)
        return output
